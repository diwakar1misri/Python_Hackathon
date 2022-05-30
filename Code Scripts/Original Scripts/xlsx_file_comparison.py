import openpyxl
import Common_Compare_data
def xlsx_comparison(files_directory,file_name_1,file_name_2):
    column_position=0
    column_name=""
    temp_line=[]
    file_1=openpyxl.load_workbook(files_directory+"\\"+file_name_1)
    file_1_sheet_name=file_1.sheetnames
    file_2 = openpyxl.load_workbook(files_directory + "\\" + file_name_2)
    file_2_sheet_name = file_2.sheetnames
    if len(file_1_sheet_name)>1 or len(file_2_sheet_name)>1:
        raise Exception("More than 1 sheet present in one of the two comparison files")
    file_1_sheet = file_1[file_1_sheet_name[0]]
    file_2_sheet = file_2[file_2_sheet_name[0]]
    file_1_max_column=file_1_sheet.max_column
    file_1_max_rows=file_1_sheet.max_row
    file_2_max_column = file_1_sheet.max_column
    file_2_max_rows = file_1_sheet.max_row
    file_1_data=[]
    file_2_data=[]
    if file_1_max_column != file_2_max_column:
        raise Exception("Header Count do not match")
    else:
        header=[]
        for itr in range(1,file_1_max_column+1):
            if file_1_sheet.cell(row=1,column=itr).value!=file_2_sheet.cell(row=1,column=itr).value:
                raise Exception("Header values do not match")
            else:
                header.append(file_1_sheet.cell(row=1,column=itr).value)
        choice=int(input("Enter:\n1.For Column Position\n2.For Column Name\n"))
        if choice==1:
            column_position=int(input("Enter the column position\n"))
        elif choice==2:
            column_name=input("Enter the column name\n")
            column_position=header.index(column_name)+1
        for outer_itr in range(2,file_1_max_rows+1):
            key=file_1_sheet.cell(row=outer_itr,column=column_position).value
            for itr in range(1,file_1_max_column+1):
                temp_line.append(file_1_sheet.cell(row=outer_itr,column=itr).value)
            file_1_data.append([temp_line[:],key,0,0])
            temp_line.clear()
#-------------------------------------------------------------------------------------------------------
        for outer_itr in range(2,file_2_max_rows+1):
            key=file_2_sheet.cell(row=outer_itr,column=column_position).value
            for itr in range(1,file_2_max_column+1):
                temp_line.append(file_2_sheet.cell(row=outer_itr,column=itr).value)
            file_2_data.append([temp_line[:],key,0,0])
            temp_line.clear()
        file_1.close()
        file_2.close()
        Common_Compare_data.compare_data(file_1_data,file_2_data)
